import json
from openpyxl import Workbook
from openpyxl.styles import Font

# Класс Seat представляет одно место в вагоне
class Seat:
    def __init__(self, number, seat_type, comfort_class, reserved=False):
        self.number = number           # Номер места
        self.seat_type = seat_type     # Тип: "нижнее" или "верхнее"
        self.comfort_class = comfort_class  # Класс: "купейное" или "плацкартное"
        self.reserved = reserved       # Статус бронирования

    def __eq__(self, other):
        # Два места считаются одинаковыми если совпадает класс комфорта и тип места
        return (self.comfort_class == other.comfort_class and 
                self.seat_type == other.seat_type)

    def to_dict(self):
        # Преобразует объект места в словарь для сохранения в JSON
        return self.__dict__

    @classmethod
    def from_dict(cls, data):
        # Создает объект места из словаря (при загрузке из файла)
        return cls(**data)

# Класс Carriage представляет железнодорожный вагон
class Carriage:
    def __init__(self, number, carriage_type):
        self.number = number          # Номер вагона
        self.carriage_type = carriage_type  # Тип: "купейный" или "плацкартный"
        self.seats = []               # Список мест в вагоне

    def add_seat(self, seat):
        # Добавляет место в вагон
        self.seats.append(seat)

    def __eq__(self, other):
        # Два вагона равны если совпадают номера и типы
        return self.number == other.number and self.carriage_type == other.carriage_type

    def to_dict(self):
        # Преобразует вагон и все его места в словарь
        return {
            'number': self.number,
            'carriage_type': self.carriage_type,
            'seats': [seat.to_dict() for seat in self.seats]  # Сохраняем все места
        }

    @classmethod
    def from_dict(cls, data):
        # Восстанавливает вагон из словаря
        carriage = cls(data['number'], data['carriage_type'])
        # Восстанавливаем все места вагона
        carriage.seats = [Seat.from_dict(seat_data) for seat_data in data['seats']]
        return carriage

# Класс Locomotive представляет локомотив поезда
class Locomotive:
    def __init__(self, serial_number, power):
        self.serial_number = serial_number  # Серийный номер
        self.power = power                  # Мощность локомотива

    def __eq__(self, other):
        # Два локомотива равны по серийному номеру
        return self.serial_number == other.serial_number

    def to_dict(self):
        return self.__dict__

    @classmethod
    def from_dict(cls, data):
        return cls(**data)

# Основной класс Train представляет весь железнодорожный состав
class Train:
    def __init__(self, number, route):
        self.number = number      # Номер поезда
        self.route = route        # Маршрут следования
        self.locomotive = None    # Локомотив поезда
        self.carriages = []       # Список вагонов в составе

    def set_locomotive(self, locomotive):
        # Устанавливает локомотив для поезда
        self.locomotive = locomotive

    def add_carriage(self, carriage):
        # Добавляет вагон в состав, если его еще нет
        if carriage not in self.carriages:
            self.carriages.append(carriage)

    def remove_carriage(self, carriage):
        # Удаляет вагон из состава
        if carriage in self.carriages:
            self.carriages.remove(carriage)

    def __eq__(self, other):
        # Два поезда равны по номеру
        return self.number == other.number

    def save_to_file(self, filename="train_data.txt"):
        # Сохраняет весь состав поезда в JSON файл
        data = {
            'number': self.number,
            'route': self.route,
            'locomotive': self.locomotive.to_dict() if self.locomotive else None,
            'carriages': [carriage.to_dict() for carriage in self.carriages]
        }
        # Записываем данные в файл с русской кодировкой
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    @classmethod
    def load_from_file(cls, filename="train_data.txt"):
        # Загружает состав поезда из JSON файла
        with open(filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Создаем объект поезда
        train = cls(data['number'], data['route'])
        # Восстанавливаем локомотив если он был
        if data['locomotive']:
            train.set_locomotive(Locomotive.from_dict(data['locomotive']))
        # Восстанавливаем все вагоны
        train.carriages = [Carriage.from_dict(carriage_data) for carriage_data in data['carriages']]
        return train

    def create_excel_report(self, filename="train_report.xlsx"):
        # Создает детальный отчет в Excel о всех местах в поезде
        wb = Workbook()
        ws = wb.active
        ws.title = f"Train {self.number}"  # Название листа

        # Создаем заголовки таблицы
        headers = ['Вагон', 'Тип вагона', 'Место', 'Тип места', 'Класс', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)  # Жирный шрифт для заголовков

        # Заполняем таблицу данными о местах
        row = 2
        for carriage in self.carriages:
            for seat in carriage.seats:
                # Записываем информацию о каждом месте
                ws.cell(row=row, column=1, value=carriage.number)
                ws.cell(row=row, column=2, value=carriage.carriage_type)
                ws.cell(row=row, column=3, value=seat.number)
                ws.cell(row=row, column=4, value=seat.seat_type)
                ws.cell(row=row, column=5, value=seat.comfort_class)
                ws.cell(row=row, column=6, value="Забронировано" if seat.reserved else "Свободно")
                row += 1

        # Автоматически настраиваем ширину колонок по содержимому
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column[0].column_letter].width = max_length + 2

        wb.save(filename)
        print(f"Отчет сохранен: {filename}")

# Демонстрация работы программы
if __name__ == "__main__":
    # Создаем тестовые данные
    
    # Создаем места разных типов
    seat1 = Seat(1, "нижнее", "купейное")
    seat2 = Seat(2, "верхнее", "купейное") 
    seat3 = Seat(1, "нижнее", "плацкартное")

    # Создаем и заполняем вагоны местами
    carriage1 = Carriage(10, "купейный")
    carriage1.add_seat(seat1)
    carriage1.add_seat(seat2)

    carriage2 = Carriage(11, "плацкартный")
    carriage2.add_seat(seat3)

    # Создаем локомотив
    locomotive = Locomotive("ТЭП-70-1234", 4000)

    # Формируем состав поезда
    train = Train("045А", "Москва - Санкт-Петербург")
    train.set_locomotive(locomotive)
    train.add_carriage(carriage1)
    train.add_carriage(carriage2)

    # Тестируем методы сравнения
    print("Сравнение мест (одинаковый класс, разный тип):", seat1 == seat2)
    print("Сравнение мест (разный класс):", seat1 == seat3)

    # Сохраняем данные в файл и загружаем обратно
    train.save_to_file()
    loaded_train = Train.load_from_file()

    # Генерируем Excel отчет о поезде
    loaded_train.create_excel_report()
